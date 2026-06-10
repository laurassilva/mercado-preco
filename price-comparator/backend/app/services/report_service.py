import io
from decimal import Decimal
from datetime import datetime

from app.schemas.product import SearchResponse


def generate_pdf(search: SearchResponse) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
    elements.append(Paragraph("Comparação de Preços", title_style))
    elements.append(Paragraph(f"Produto: <b>{search.query}</b>  |  Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                               styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))

    if search.avg_price:
        summary_data = [
            ["Menor Preço", "Maior Preço", "Preço Médio", "Total de Resultados"],
            [
                f"R$ {search.results[0].price:.2f} ({search.results[0].market_name})" if search.results else "-",
                f"R$ {search.results[-1].price:.2f} ({search.results[-1].market_name})" if search.results else "-",
                f"R$ {search.avg_price:.2f}",
                str(search.total),
            ],
        ]
        summary_table = Table(summary_data, colWidths=[6*cm, 6*cm, 5*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f0f9ff")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f9ff")]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.5*cm))

    headers = ["Mercado", "Produto", "Marca", "Quantidade", "Preço (R$)", "Diferença (R$)", "Link"]
    data = [headers]
    for r in search.results:
        row_color = colors.HexColor("#dcfce7") if r.is_cheapest else colors.white
        data.append([
            r.market_name,
            r.product_name[:40],
            r.brand or "-",
            r.quantity or "-",
            f"{r.price:.2f}",
            f"+{r.difference:.2f}" if r.difference else "0,00",
            (r.product_url or "")[:30],
        ])

    col_widths = [4.5*cm, 6*cm, 3.5*cm, 3*cm, 3*cm, 3.5*cm, 5*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ALIGN", (4, 0), (5, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])
    if len(data) > 1:
        style.add("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#dcfce7"))
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    return buffer.getvalue()


def generate_excel(search: SearchResponse) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparação de Preços"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Comparação de Preços – {search.query}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(size=10, italic=True)

    headers = ["Mercado", "Produto", "Marca", "Quantidade", "Preço (R$)", "Diferença (R$)", "Link"]
    header_fill = PatternFill("solid", fgColor="1e40af")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    green_fill = PatternFill("solid", fgColor="dcfce7")
    alt_fill = PatternFill("solid", fgColor="f8fafc")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for row_idx, r in enumerate(search.results, 5):
        fill = green_fill if r.is_cheapest else (PatternFill("solid", fgColor="ffffff") if row_idx % 2 == 1 else alt_fill)
        values = [
            r.market_name, r.product_name, r.brand or "",
            r.quantity or "", float(r.price),
            float(r.difference) if r.difference else 0.0,
            r.product_url or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            if col_idx in (5, 6):
                cell.number_format = 'R$ #,##0.00'

    col_widths_chars = [20, 40, 18, 12, 14, 14, 35]
    for i, w in enumerate(col_widths_chars, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_csv(search: SearchResponse) -> str:
    import csv, io as sio
    out = sio.StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(["Mercado", "Produto", "Marca", "Quantidade", "Preço", "Diferença", "Link"])
    for r in search.results:
        writer.writerow([
            r.market_name, r.product_name, r.brand or "",
            r.quantity or "", str(r.price).replace(".", ","),
            str(r.difference or "0").replace(".", ","),
            r.product_url or "",
        ])
    return out.getvalue()
