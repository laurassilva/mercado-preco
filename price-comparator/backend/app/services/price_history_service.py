"""Geração de PDF e Excel para histórico de preços."""
from io import BytesIO
from datetime import datetime, timezone

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm


def generate_history_pdf(query: str, period: str, products: list) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Histórico de Preços — {query}", styles["Title"]))
    elements.append(Paragraph(
        f"Período: {period} | Gerado em: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 10*mm))

    for product in products:
        elements.append(Paragraph(
            f"{product.product_name} — {product.market_name}",
            styles["Heading3"],
        ))

        data = [["Data", "Preço (R$)"]]
        for entry in product.history:
            data.append([
                entry.checked_at.strftime("%d/%m/%Y %H:%M"),
                f"R$ {entry.price:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            ])

        if len(data) > 1:
            table = Table(data, colWidths=[120, 100])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]))
            elements.append(table)

        elements.append(Spacer(1, 8*mm))

    doc.build(elements)
    return buffer.getvalue()


def generate_history_excel(query: str, period: str, products: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Preços"

    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    alt_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Histórico de Preços — {query}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Período: {period} | Gerado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC"

    row = 4
    headers = ["Mercado", "Produto", "Data", "Preço (R$)", "Categoria"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    row = 5
    for product in products:
        for entry in product.history:
            ws.cell(row=row, column=1, value=product.market_name)
            ws.cell(row=row, column=2, value=product.product_name)
            ws.cell(row=row, column=3, value=entry.checked_at.strftime("%d/%m/%Y %H:%M"))
            ws.cell(row=row, column=4, value=float(entry.price))
            ws.cell(row=row, column=4).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=5, value=product.category or "-")
            if row % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = alt_fill
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
